"""
Online Examination System — Complete Views Layer
=================================================
Covers both the student-facing interface and the custom admin dashboard.
All admin views are gated by SuperuserRequiredMixin (is_superuser=True).
All student views are gated by StudentRequiredMixin.
"""

import json
from django.contrib import messages
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Avg, Count, Sum, Q, F
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views import View
from django.views.generic import TemplateView, ListView, DetailView

from .forms import (
    StudentRegistrationForm,
    CustomLoginForm,
    StudentProfileForm,
    AdminProfileForm,
    CategoryForm,
    QuestionForm,
    ExamRequestForm,
    ExcelImportForm,
    ExamRequestReviewForm,
    ForgotPasswordForm,
    OTPVerificationForm,
    ResetPasswordForm,
    SubAdminForm,
)
from .mixins import SuperuserRequiredMixin, StudentRequiredMixin, BaseAdminRequiredMixin
from .models import (
    CustomUser,
    Category,
    Question,
    ExamPaper,
    ExamPaperQuestion,
    StudentExamResult,
    StudentAnswer,
    ExamRequest,
    Notification,
    OTP,
    SubAdminOTP,
    Certificate,
)
from .utils import import_questions_from_excel, generate_exam_paper, submit_and_evaluate


# ════════════════════════════════════════════════
#  AUTHENTICATION VIEWS
# ════════════════════════════════════════════════

class RegisterView(View):
    """Student registration — creates user with role=Student."""

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if request.user.is_superuser or request.user.role == CustomUser.Role.SUB_ADMIN:
                return redirect('admin_dashboard')
            return redirect('student_dashboard')
        response = super().dispatch(request, *args, **kwargs)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    def get(self, request):
        form = StudentRegistrationForm()
        return render(request, 'exams/register.html', {'form': form})

    def post(self, request):
        form = StudentRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            # Ensure no existing inactive user blocks this registration
            email = form.cleaned_data.get('email')
            CustomUser.objects.filter(email=email, is_active=False).delete()
            
            # Save user and activate immediately
            user = form.save()
            
            # Log them in
            login(request, user)
            messages.success(request, 'Registration complete! Welcome to Aptipro.')
            return redirect('student_dashboard')
        return render(request, 'exams/register.html', {'form': form})

class ValidateStudentFieldAPI(View):
    """API for real-time duplicate checking during student registration."""
    def get(self, request):
        field = request.GET.get('field', '').strip()
        value = request.GET.get('value', '').strip()

        if not field or not value:
            return JsonResponse({'exists': False, 'error': 'Missing field or value'}, status=400)

        # Allowed fields for duplicate checking
        allowed_fields = ['username', 'email', 'phone_number', 'admission_number', 'roll_number']
        if field not in allowed_fields:
            return JsonResponse({'exists': False, 'error': 'Invalid field'}, status=400)

        # Construct the query
        query = {field: value}
        
        # Check existence
        exists = CustomUser.objects.filter(**query).exists()
        
        return JsonResponse({'exists': exists, 'field': field, 'value': value})

class LoginView(View):
    """Login view — routes to appropriate dashboard based on role."""

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if request.user.is_superuser or request.user.role == CustomUser.Role.SUB_ADMIN:
                return redirect('admin_dashboard')
            return redirect('student_dashboard')
        response = super().dispatch(request, *args, **kwargs)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    def get(self, request):
        form = CustomLoginForm()
        return render(request, 'exams/login.html', {'form': form})

    def post(self, request):
        form = CustomLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f'Welcome back, {user.username}!')

            # Superusers (including accounts created via `python manage.py
            # createsuperuser`) ALWAYS land on the admin dashboard, regardless
            # of what the `role` field happens to be — `createsuperuser` does
            # not touch the role column, so it stays at its model default
            # ('Student'), and we must not let that mislead the redirect.
            if user.is_superuser:
                return redirect('admin_dashboard')
            # Sub-Admins (non-superuser staff role) also use the admin panel.
            if user.role == CustomUser.Role.SUB_ADMIN:
                return redirect('admin_dashboard')
            # Everyone else is a regular student.
            return redirect('student_dashboard')
        return render(request, 'exams/login.html', {'form': form})


class LogoutView(View):
    """Logout and redirect to login."""
    def get(self, request):
        logout(request)
        messages.info(request, 'You have been logged out.')
        return redirect('login')

    def post(self, request):
        logout(request)
        return redirect('login')


class ForgotPasswordView(View):
    """Initial step: user enters email to receive an OTP."""
    def get(self, request):
        form = ForgotPasswordForm()
        return render(request, 'exams/forgot_password.html', {'form': form})

    def post(self, request):
        form = ForgotPasswordForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            user = CustomUser.objects.get(email=email)
            
            # Generate 6-digit random OTP
            import random
            otp_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
            
            # Save OTP to DB
            OTP.objects.create(user=user, code=otp_code)
            
            # Send Email
            subject = f"Your {settings.DEFAULT_FROM_EMAIL.split('@')[0]} Password Reset OTP"
            # Actually use 'Aptipro' here as requested
            subject = "Aptipro Password Reset OTP"
            message = f"Hello {user.username},\n\nYour OTP for resetting your password is: {otp_code}\n\nThis OTP is valid for 10 minutes.\n\nRegards,\nAptipro Team"
            
            try:
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
                messages.success(request, f"An OTP has been sent to {email}.")
                request.session['reset_email'] = email
                return redirect('verify_otp')
            except Exception as e:
                messages.error(request, f"Failed to send email. Please check your SMTP settings. Error: {str(e)}")
                
        return render(request, 'exams/forgot_password.html', {'form': form})


class VerifyOTPView(View):
    """Step 2: user enters the OTP they received."""
    def get(self, request):
        if 'reset_email' not in request.session:
            return redirect('forgot_password')
        form = OTPVerificationForm()
        return render(request, 'exams/verify_otp.html', {'form': form})

    def post(self, request):
        if 'reset_email' not in request.session:
            return redirect('forgot_password')
        
        email = request.session['reset_email']
        user = get_object_or_404(CustomUser, email=email)
        form = OTPVerificationForm(request.POST)
        
        if form.is_valid():
            code = form.cleaned_data['code']
            otp_obj = OTP.objects.filter(user=user, code=code, is_verified=False).last()
            
            if otp_obj and not otp_obj.is_expired():
                otp_obj.is_verified = True
                otp_obj.save()
                messages.success(request, "OTP verified successfully. You can now reset your password.")
                return redirect('reset_password')
            else:
                messages.error(request, "Invalid or expired OTP.")
                
        return render(request, 'exams/verify_otp.html', {'form': form})


class ResetPasswordView(View):
    """Step 3: user sets a new password."""
    def get(self, request):
        if 'reset_email' not in request.session:
            return redirect('forgot_password')
        
        email = request.session['reset_email']
        user = get_object_or_404(CustomUser, email=email)
        # Ensure latest OTP is verified
        otp_obj = OTP.objects.filter(user=user, is_verified=True).last()
        if not otp_obj:
            messages.error(request, "Please verify your OTP first.")
            return redirect('verify_otp')

        form = ResetPasswordForm()
        return render(request, 'exams/reset_password.html', {'form': form})

    def post(self, request):
        if 'reset_email' not in request.session:
            return redirect('forgot_password')
        
        email = request.session['reset_email']
        user = get_object_or_404(CustomUser, email=email)
        otp_obj = OTP.objects.filter(user=user, is_verified=True).last()
        if not otp_obj:
            return redirect('verify_otp')

        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['password']
            user.set_password(new_password)
            user.save()
            
            # For security, invalidate the session and session hash
            update_session_auth_hash(request, user)
            
            # Clean up session
            del request.session['reset_email']
            
            messages.success(request, "Your password has been reset successfully. You can now log in.")
            return redirect('login')
            
        return render(request, 'exams/reset_password.html', {'form': form})


# ════════════════════════════════════════════════
#  SUB-ADMIN AUTHENTICATION (FORGOT PASSWORD)
# ════════════════════════════════════════════════

class SubAdminForgotPasswordView(View):
    """Sub-Admins enter their email, OTP goes ONLY to ADMIN."""
    def get(self, request):
        form = ForgotPasswordForm()
        return render(request, 'exams/subadmin/forgot_password.html', {'form': form})

    def post(self, request):
        form = ForgotPasswordForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            user = CustomUser.objects.filter(email=email, role=CustomUser.Role.SUB_ADMIN).first()
            
            if user:
                # Generate OTP
                import random
                otp_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
                
                # Save OTP
                SubAdminOTP.objects.filter(user=user).delete()
                SubAdminOTP.objects.create(
                    user=user, 
                    otp=otp_code,
                    expires_at=timezone.now() + timezone.timedelta(minutes=5)
                )
                
                # Send Email ONLY to Admin
                subject = "Sub-Admin Password Reset Request"
                message = f"Hello Admin,\n\nSub-Admin '{user.username}' has requested a password reset.\n\nPlease share this OTP with them manually: {otp_code}\n\nRegards,\nAptipro System"
                
                try:
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [settings.ADMIN_EMAIL],
                        fail_silently=False,
                    )
                    request.session['subadmin_reset_email'] = email
                    messages.success(request, "A reset request has been sent to the Administrator. Please contact them for your OTP.")
                    return redirect('subadmin_verify_otp')
                except Exception as e:
                    messages.error(request, f"Failed to notify admin: {str(e)}")
            else:
                messages.error(request, "No sub-admin found with this email.")
                
        return render(request, 'exams/subadmin/forgot_password.html', {'form': form})


class SubAdminVerifyOTPView(View):
    """Step 2 for Sub-Admin password reset."""
    def get(self, request):
        if 'subadmin_reset_email' not in request.session:
            return redirect('subadmin_forgot_password')
        form = OTPVerificationForm()
        return render(request, 'exams/subadmin/verify_otp.html', {'form': form})

    def post(self, request):
        if 'subadmin_reset_email' not in request.session:
            return redirect('subadmin_forgot_password')
        
        email = request.session['subadmin_reset_email']
        user = get_object_or_404(CustomUser, email=email, role=CustomUser.Role.SUB_ADMIN)
        form = OTPVerificationForm(request.POST)
        
        if form.is_valid():
            code = form.cleaned_data['code']
            otp_obj = SubAdminOTP.objects.filter(user=user, otp=code, is_verified=False).last()
            
            if otp_obj and not otp_obj.is_expired():
                otp_obj.is_verified = True
                otp_obj.save()
                messages.success(request, "OTP verified. You can now set a new password.")
                return redirect('subadmin_reset_password')
            else:
                messages.error(request, "Invalid or expired OTP.")
                
        return render(request, 'exams/subadmin/verify_otp.html', {'form': form})


class SubAdminResetPasswordView(View):
    """Step 3 for Sub-Admin password reset."""
    def get(self, request):
        if 'subadmin_reset_email' not in request.session:
            return redirect('subadmin_forgot_password')
        
        email = request.session['subadmin_reset_email']
        user = get_object_or_404(CustomUser, email=email, role=CustomUser.Role.SUB_ADMIN)
        otp_obj = SubAdminOTP.objects.filter(user=user, is_verified=True).last()
        if not otp_obj:
            return redirect('subadmin_verify_otp')

        form = ResetPasswordForm()
        return render(request, 'exams/subadmin/reset_password.html', {'form': form})

    def post(self, request):
        if 'subadmin_reset_email' not in request.session:
            return redirect('subadmin_forgot_password')
        
        email = request.session['subadmin_reset_email']
        user = get_object_or_404(CustomUser, email=email, role=CustomUser.Role.SUB_ADMIN)
        otp_obj = SubAdminOTP.objects.filter(user=user, is_verified=True).last()
        if not otp_obj:
            return redirect('subadmin_verify_otp')

        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['password']
            user.set_password(new_password)
            user.raw_password = new_password # Update raw password for admin visibility
            user.save()
            
            # Clean up
            otp_obj.delete()
            del request.session['subadmin_reset_email']
            
            messages.success(request, "Password reset successful. You can now log in.")
            return redirect('login')
            
        return render(request, 'exams/subadmin/reset_password.html', {'form': form})


@login_required
def dashboard_redirect(request):
    """Route /dashboard/ to the correct interface based on role.

    Superusers (Admins) and Sub-Admins are sent to the admin dashboard;
    every other authenticated user goes to the student dashboard. This
    mirrors the routing logic in ``LoginView.dispatch`` / ``LoginView.post``
    so that the post-login landing page is consistent regardless of whether
    the user arrives via the login form or via ``@login_required``'s
    ``LOGIN_REDIRECT_URL``.
    """
    user = request.user
    if user.is_superuser or user.role == CustomUser.Role.SUB_ADMIN:
        return redirect('admin_dashboard')
    return redirect('student_dashboard')


# ════════════════════════════════════════════════
#  STUDENT VIEWS
# ════════════════════════════════════════════════

class StudentDashboardView(StudentRequiredMixin, TemplateView):
    """Student home page — shows exam status, notifications, request CTA."""
    template_name = 'exams/student/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user

        # Exam requests
        ctx['pending_requests'] = ExamRequest.objects.filter(
            student=user, status='Pending'
        ).count()
        ctx['approved_requests'] = ExamRequest.objects.filter(
            student=user, status='Approved'
        )
        ctx['recent_requests'] = ExamRequest.objects.filter(
            student=user
        ).order_by('-requested_at')[:5]

        # Active exams (approved but not yet taken)
        taken_paper_ids = StudentExamResult.objects.filter(
            student=user
        ).values_list('exam_paper_id', flat=True)
        ctx['available_exams'] = ExamPaper.objects.filter(
            student=user,
            category__is_active=True
        ).exclude(id__in=taken_paper_ids)

        # Completed exams count (students see only "Test Completed")
        ctx['completed_exams'] = StudentExamResult.objects.filter(
            student=user, status__in=['Submitted', 'Evaluated']
        ).count()

        # Unread notifications
        ctx['unread_count'] = Notification.unread_count(user) or 0
        ctx['notifications'] = Notification.objects.filter(
            recipient=user
        ).order_by('-created_at')[:10]

        # Active categories for the request modal
        ctx['active_categories'] = Category.objects.filter(is_active=True).order_by('name')

        return ctx


class StudentProfileView(StudentRequiredMixin, View):
    """View and update student profile."""

    def get(self, request):
        form = StudentProfileForm(instance=request.user)
        return render(request, 'exams/student/profile.html', {'form': form})

    def post(self, request):
        form = StudentProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('student_profile')
        else:
            messages.error(request, 'Please correct the errors below.')
        return render(request, 'exams/student/profile.html', {'form': form})


class AdminProfileView(SuperuserRequiredMixin, View):
    """View and update admin profile."""

    def get(self, request):
        form = AdminProfileForm(instance=request.user)
        return render(request, 'exams/admin/profile.html', {'form': form})

    def post(self, request):
        form = AdminProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('admin_profile')
        else:
            messages.error(request, 'Please correct the errors below.')
        return render(request, 'exams/admin/profile.html', {'form': form})


class SubAdminProfileView(BaseAdminRequiredMixin, View):
    """View and update sub-admin profile."""

    def get(self, request):
        form = AdminProfileForm(instance=request.user)
        return render(request, 'exams/subadmin/profile.html', {'form': form})

    def post(self, request):
        form = AdminProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('subadmin_profile')
        else:
            messages.error(request, 'Please correct the errors below.')
        return render(request, 'exams/subadmin/profile.html', {'form': form})


class RequestExamView(StudentRequiredMixin, View):
    """Student submits an exam access request for a specific category."""

    def post(self, request):
        form = ExamRequestForm(request.POST)
        if form.is_valid():
            category = form.cleaned_data['category']
            user = request.user
            
            # 1. Block if there is already a Pending request for this category
            if ExamRequest.objects.filter(student=user, category=category, status='Pending').exists():
                messages.warning(request, f'You already have a pending request for {category.name}.')
                return redirect('student_dashboard')

            # 2. If they have an Approved request but no result yet (Unattended)
            # We still allow a re-request if they feel they need a new approval/paper
            # per user requirement: "must again send a request"

            # 3. Check if they have already passed — per spec, re-exam is for those who "do not pass"
            # But the user also mentions "does not attend".
            last_result = StudentExamResult.objects.filter(
                student=user, exam_paper__category=category
            ).order_by('-submitted_at').first()
            
            if last_result and last_result.status == 'Evaluated' and last_result.is_passed():
                # They passed already. Still allow? User's "does not pass" implies failure.
                # I'll allow but show a specific info message.
                pass 
                
            ExamRequest.objects.create(
                student=user,
                category=category
            )
            messages.success(request, f'Request for {category.name} submitted! Please wait for admin approval.')
        else:
            messages.error(request, 'Please select a valid category.')
            
        return redirect('student_dashboard')


class StudentCategoriesView(StudentRequiredMixin, ListView):
    """List all categories for students with contextual actions (Request/Retest/Start)."""
    template_name = 'exams/student/categories.html'
    context_object_name = 'categories'
    
    def get_queryset(self):
        return Category.objects.filter(is_active=True).annotate(
            q_count=Count('questions', filter=Q(questions__is_active=True))
        ).order_by('name')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Pending requests
        ctx['pending_cat_ids'] = list(ExamRequest.objects.filter(
            student=user, status='Pending'
        ).values_list('category_id', flat=True))
        
        # Approved but not taken (Access Granted)
        # We define "not taken" as having an active paper without a submitted result
        untaken_papers = ExamPaper.objects.filter(
            student=user
        ).exclude(
            result__status__in=['Submitted', 'Evaluated']
        ).values('category_id', 'id')
        
        # Build a map for easy lookup
        paper_map = {p['category_id']: p['id'] for p in untaken_papers}

        # Inject paper_id into categories for the template
        for cat in ctx['categories']:
            cat.paper_id = paper_map.get(cat.id)
        
        # Categories already attempted (to show "Retest" vs "Request")
        completed_cat_ids = StudentExamResult.objects.filter(
            student=user,
            status__in=['Submitted', 'Evaluated']
        ).values_list('exam_paper__category_id', flat=True).distinct()
        ctx['completed_cat_ids'] = list(completed_cat_ids)
        
        # Identify failed categories explicitly
        # This is a bit complex in one line, so we do it in a loop or filtered list
        results = StudentExamResult.objects.filter(student=user, status='Evaluated')
        failed_cat_ids = []
        for r in results:
            if not r.is_passed():
                failed_cat_ids.append(r.exam_paper.category_id)
        ctx['failed_cat_ids'] = list(set(failed_cat_ids))

        return ctx


class TakeExamView(StudentRequiredMixin, View):
    """Student takes an exam — shows question paper."""

    def get(self, request, paper_id):
        paper = get_object_or_404(ExamPaper, id=paper_id, student=request.user)

        # Check if already submitted or evaluated
        if StudentExamResult.objects.filter(
            exam_paper=paper, 
            status__in=[StudentExamResult.Status.SUBMITTED, StudentExamResult.Status.EVALUATED]
        ).exists():
            messages.warning(request, 'You have already completed this exam.')
            return redirect('student_dashboard')

        # Create result record (In Progress)
        result, created = StudentExamResult.objects.get_or_create(
            student=request.user,
            exam_paper=paper,
            defaults={
                'total_marks_possible': paper.total_marks,
                'status': StudentExamResult.Status.IN_PROGRESS,
            }
        )

        questions = paper.paper_questions.select_related('question').order_by('order')
        response = render(request, 'exams/student/take_exam.html', {
            'paper': paper,
            'questions': questions,
            'result': result,
        })
        # Prevent caching test questions
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    def post(self, request, paper_id):
        paper = get_object_or_404(ExamPaper, id=paper_id, student=request.user)
        result = get_object_or_404(
            StudentExamResult,
            exam_paper=paper,
            student=request.user,
            status=StudentExamResult.Status.IN_PROGRESS,
        )

        # Save all answers
        questions = paper.paper_questions.select_related('question').order_by('order')
        answers_to_create = []
        for pq in questions:
            answer_text = request.POST.get(f'answer_{pq.question.id}', '').strip()
            answers_to_create.append(
                StudentAnswer(
                    result=result,
                    question=pq.question,
                    student_answer=answer_text,
                )
            )

        if answers_to_create:
            StudentAnswer.objects.bulk_create(answers_to_create, ignore_conflicts=True)

        # Submit and auto-evaluate
        result.submitted_at = timezone.now()
        result.status = StudentExamResult.Status.SUBMITTED
        result.save(update_fields=['submitted_at', 'status'])
        result.auto_evaluate()

        messages.success(request, 'Test Completed! Your responses have been submitted.')
        return redirect('exam_complete')


class ExamCompleteView(StudentRequiredMixin, TemplateView):
    """Shows "Test Completed" message with optional congratulations popup."""
    template_name = 'exams/student/exam_complete.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Get the latest result for this student
        latest_result = StudentExamResult.objects.filter(
            student=user
        ).select_related('exam_paper', 'exam_paper__category').order_by('-submitted_at').first()
        
        if latest_result and latest_result.status == 'Evaluated':
            percentage = latest_result.percentage()
            if percentage >= 85:
                ctx['show_congratulations'] = True
                ctx['result'] = latest_result
                ctx['percentage'] = percentage
                ctx['marks'] = latest_result.total_marks_obtained
                ctx['total'] = latest_result.total_marks_possible
                ctx['exam_name'] = latest_result.exam_paper.category.name
                ctx['category_name'] = latest_result.exam_paper.category.name
                # Domain is typically a field in Category, if not we use Category name
                ctx['domain_name'] = getattr(latest_result.exam_paper.category, 'domain', 'Aptitude')
        
        return ctx

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        # Prevent back-navigation to the test from here
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response


class StudentHistoryView(StudentRequiredMixin, ListView):
    """Student's past exam attempts & performance."""
    template_name = 'exams/student/history.html'
    context_object_name = 'results'

    def get_queryset(self):
        return StudentExamResult.objects.filter(
            student=self.request.user,
            status__in=['Submitted', 'Evaluated']
        ).select_related('exam_paper__category').order_by('-submitted_at')


class StudentNotificationsView(StudentRequiredMixin, ListView):
    """Student notification list."""
    template_name = 'exams/student/notifications.html'
    context_object_name = 'notifications'
    paginate_by = 20

    def get_queryset(self):
        return Notification.objects.filter(
            recipient=self.request.user
        ).order_by('-created_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['unread_count'] = Notification.unread_count(self.request.user)
        return ctx


class MarkNotificationReadView(LoginRequiredMixin, View):
    """AJAX endpoint to mark a notification as read."""

    def post(self, request, notification_id):
        notif = get_object_or_404(
            Notification, id=notification_id, recipient=request.user
        )
        notif.mark_as_read()
        return JsonResponse({'status': 'ok'})


class MarkAllNotificationsReadView(LoginRequiredMixin, View):
    """Mark all notifications read for the current user."""

    def post(self, request):
        Notification.mark_all_read(request.user)
        messages.success(request, 'All notifications marked as read.')
        if request.user.is_superuser:
            return redirect('admin_notifications')
        return redirect('student_notifications')


class FetchNotificationsView(LoginRequiredMixin, View):
    """JSON API to fetch latest unread notifications for the dropdown."""
    def get(self, request):
        count = Notification.unread_count(request.user)
        latest = Notification.objects.filter(
            recipient=request.user, 
            is_read=False
        ).order_by('-created_at')[:5]
        
        data = {
            'count': count,
            'notifications': [
                {
                    'id': str(n.id),
                    'title': n.title,
                    'message': n.message,
                    'created_at': n.created_at.strftime('%I:%M %p'),
                    'type': n.notification_type
                } for n in latest
            ]
        }
        return JsonResponse(data)


class DeleteNotificationView(LoginRequiredMixin, View):
    """Delete a specific notification."""
    def post(self, request, notification_id):
        notif = get_object_or_404(
            Notification, id=notification_id, recipient=request.user
        )
        notif.delete()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'ok'})
            
        messages.success(request, 'Notification deleted.')
        if request.user.is_superuser:
            return redirect('admin_notifications')
        return redirect('student_notifications')


# ════════════════════════════════════════════════
#  SUB-ADMIN MANAGEMENT (SUPERUSER ONLY)
# ════════════════════════════════════════════════

class AdminSubAdminsView(SuperuserRequiredMixin, ListView):
    """List all Sub-Admins."""
    template_name = 'exams/admin/subadmins.html'
    context_object_name = 'subadmins'
    paginate_by = 10
    
    def get_queryset(self):
        qs = CustomUser.objects.filter(role=CustomUser.Role.SUB_ADMIN).order_by('-date_joined')
        search = self.request.GET.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form'] = SubAdminForm()
        ctx['search'] = self.request.GET.get('search', '')
        return ctx


class AdminAddSubAdminView(SuperuserRequiredMixin, View):
    """Add a new Sub-Admin."""
    def post(self, request):
        form = SubAdminForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Sub-Admin created successfully.")
        else:
            # Handle errors (e.g. username taken)
            error_msg = list(form.errors.values())[0][0]
            messages.error(request, f"Error: {error_msg}")
        return redirect('admin_subadmins')


class AdminEditSubAdminView(SuperuserRequiredMixin, View):
    """Edit an existing Sub-Admin."""
    def post(self, request, pk):
        subadmin = get_object_or_404(CustomUser, pk=pk, role=CustomUser.Role.SUB_ADMIN)
        form = SubAdminForm(request.POST, instance=subadmin)
        if form.is_valid():
            form.save()
            messages.success(request, "Sub-Admin updated successfully.")
        else:
            error_msg = list(form.errors.values())[0][0]
            messages.error(request, f"Error: {error_msg}")
        return redirect('admin_subadmins')


class AdminDeleteSubAdminView(SuperuserRequiredMixin, View):
    """Delete a Sub-Admin."""
    def post(self, request, pk):
        subadmin = get_object_or_404(CustomUser, pk=pk, role=CustomUser.Role.SUB_ADMIN)
        username = subadmin.username
        subadmin.delete()
        messages.success(request, f"Sub-Admin '{username}' deleted successfully.")
        return redirect('admin_subadmins')


# ════════════════════════════════════════════════
#  SUB-ADMIN API ENDPOINTS
# ════════════════════════════════════════════════

class SubAdminOTPAPI(View):
    """API for sub-admin forgot password OTP."""
    def post(self, request):
        email = request.POST.get('email')
        user = CustomUser.objects.filter(email=email, role=CustomUser.Role.SUB_ADMIN).first()
        
        if not user:
            return JsonResponse({'status': 'error', 'message': 'Sub-Admin not found.'}, status=404)

        import random
        otp_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        
        SubAdminOTP.objects.filter(user=user).delete()
        SubAdminOTP.objects.create(
            user=user, 
            otp=otp_code,
            expires_at=timezone.now() + timezone.timedelta(minutes=5)
        )
        
        subject = "Sub-Admin Password Reset OTP"
        message = f"OTP for {user.username}: {otp_code}"
        
        try:
            send_mail(
                subject, message, settings.DEFAULT_FROM_EMAIL, [settings.ADMIN_EMAIL]
            )
            return JsonResponse({'status': 'success', 'message': 'OTP sent to Admin.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# ════════════════════════════════════════════════
#  ADMIN DASHBOARD VIEWS (SUPERUSER ONLY)
# ════════════════════════════════════════════════

class AdminDashboardView(BaseAdminRequiredMixin, TemplateView):
    """Admin dashboard home — overview analytics."""
    template_name = 'exams/admin/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # Counts
        ctx['total_students'] = CustomUser.objects.filter(role='Student').count()
        ctx['total_questions'] = Question.objects.filter(is_active=True).count()
        ctx['total_exams'] = StudentExamResult.objects.count()
        ctx['pending_requests'] = ExamRequest.objects.filter(status='Pending').count()

        # Students who registered but never attended exam
        attended_student_ids = StudentExamResult.objects.values_list('student_id', flat=True).distinct()
        ctx['never_attended_count'] = CustomUser.objects.filter(role='Student').exclude(id__in=attended_student_ids).count()
        ctx['attended_count'] = len(attended_student_ids)

        # Recent requests
        ctx['recent_requests'] = ExamRequest.objects.select_related(
            'student'
        ).order_by('-requested_at')[:5]

        # Recent results
        ctx['recent_results'] = StudentExamResult.objects.select_related(
            'student', 'exam_paper'
        ).filter(status='Evaluated').order_by('-completed_at')[:5]

        # Pass/fail stats
        evaluated = StudentExamResult.objects.filter(status='Evaluated')
        total_evaluated = evaluated.count()
        if total_evaluated > 0:
            passed = sum(1 for r in evaluated.only(
                'total_marks_obtained', 'total_marks_possible'
            ) if r.is_passed())
            ctx['pass_rate'] = round((passed / total_evaluated) * 100, 1)
            ctx['fail_rate'] = round(100 - ctx['pass_rate'], 1)
        else:
            ctx['pass_rate'] = 0
            ctx['fail_rate'] = 0

        # Average score
        avg = evaluated.aggregate(avg_score=Avg('total_marks_obtained'))
        ctx['avg_score'] = round(avg['avg_score'] or 0, 1)

        # Unread notifications
        ctx['unread_count'] = Notification.unread_count(self.request.user)

        # Question categories for chart
        categories = Question.objects.filter(is_active=True).values(
            'category__name'
        ).annotate(count=Count('id')).order_by('-count')[:10]
        ctx['category_labels'] = json.dumps([c['category__name'] for c in categories])
        ctx['category_data'] = json.dumps([c['count'] for c in categories])

        # Difficulty distribution
        difficulties = Question.objects.filter(is_active=True).values(
            'difficulty'
        ).annotate(count=Count('id'))
        diff_dict = {d['difficulty']: d['count'] for d in difficulties}
        ctx['difficulty_data'] = json.dumps([
            diff_dict.get('Easy', 0),
            diff_dict.get('Medium', 0),
            diff_dict.get('Hard', 0),
        ])

        return ctx


class AdminStudentsView(BaseAdminRequiredMixin, ListView):
    """View all students with full profile details."""
    template_name = 'exams/admin/students.html'
    context_object_name = 'students'
    paginate_by = 20

    def get_queryset(self):
        qs = CustomUser.objects.filter(role='Student').order_by('-date_joined')
        search = self.request.GET.get('search', '').strip()
        min_pct = self.request.GET.get('min_percentage', '').strip()

        if search:
            qs = qs.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(institution__icontains=search)
            )

        if min_pct:
            try:
                val = float(min_pct)
                from django.db.models.functions import Cast
                from django.db.models import FloatField
                valid_student_ids = StudentExamResult.objects.filter(
                    total_marks_possible__gt=0
                ).annotate(
                    calc_pct=Cast(F('total_marks_obtained'), FloatField()) * 100 / Cast(F('total_marks_possible'), FloatField())
                ).filter(calc_pct__gte=val).values_list('student_id', flat=True)
                qs = qs.filter(id__in=valid_student_ids)
            except ValueError:
                pass

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search'] = self.request.GET.get('search', '')
        ctx['min_percentage'] = self.request.GET.get('min_percentage', '')
        ctx['total_students'] = self.get_queryset().count()
        return ctx


class AdminStudentDetailView(BaseAdminRequiredMixin, DetailView):
    """Full student profile visible to admin."""
    template_name = 'exams/admin/student_detail.html'
    context_object_name = 'student'

    def get_queryset(self):
        return CustomUser.objects.filter(role='Student')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        student = self.object

        # Exam history with results
        ctx['exam_results'] = StudentExamResult.objects.filter(
            student=student
        ).select_related('exam_paper', 'exam_paper__category').order_by('-started_at')

        # Request history
        ctx['exam_requests'] = ExamRequest.objects.filter(
            student=student
        ).order_by('-requested_at')

        # Certificate stats
        ctx['certificates'] = Certificate.objects.filter(student=student).select_related('category')
        ctx['certificate_count'] = ctx['certificates'].count()

        return ctx


class AdminDeleteStudentView(SuperuserRequiredMixin, View):
    """Safely delete a student and all related records (cascade handled by DB)."""
    def post(self, request, pk):
        student = get_object_or_404(CustomUser, pk=pk, role='Student')
        username = student.username
        student.delete()
        messages.success(request, f'Student "{username}" has been removed from the platform.')
        return redirect('admin_students')


class AdminQuestionsView(BaseAdminRequiredMixin, ListView):
    """Question bank management."""
    template_name = 'exams/admin/questions.html'
    context_object_name = 'questions'
    paginate_by = 25

    def get_queryset(self):
        qs = Question.objects.select_related('category').filter(is_active=True).order_by('-created_at')
        category = self.request.GET.get('category', '').strip()
        difficulty = self.request.GET.get('difficulty', '').strip()
        q_type = self.request.GET.get('type', '').strip()
        search = self.request.GET.get('search', '').strip()

        if category:
            qs = qs.filter(category=category)
        if difficulty:
            qs = qs.filter(difficulty=difficulty)
        if q_type:
            qs = qs.filter(question_type=q_type)
        if search:
            qs = qs.filter(question_text__icontains=search)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['categories'] = Category.objects.all().order_by('name')
        ctx['difficulties'] = Question.Difficulty.choices
        ctx['question_types'] = Question.QuestionType.choices
        ctx['current_category'] = self.request.GET.get('category', '')
        ctx['current_difficulty'] = self.request.GET.get('difficulty', '')
        ctx['current_type'] = self.request.GET.get('type', '')
        ctx['search'] = self.request.GET.get('search', '')
        ctx['total_questions'] = Question.objects.filter(is_active=True).count()
        return ctx


class AdminCategoryDetailView(BaseAdminRequiredMixin, ListView):
    """View all questions within a specific category."""
    template_name = 'exams/admin/category_detail.html'
    context_object_name = 'questions'
    paginate_by = 25

    def get_queryset(self):
        self.category = get_object_or_404(Category, pk=self.kwargs['pk'])
        qs = Question.objects.filter(category=self.category, is_active=True).order_by('-created_at')
        
        # Simple search within category
        search = self.request.GET.get('search', '').strip()
        if search:
            qs = qs.filter(question_text__icontains=search)
            
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['category'] = self.category
        ctx['search'] = self.request.GET.get('search', '')
        return ctx


class AdminAddQuestionView(BaseAdminRequiredMixin, View):
    """Add individual question, optionally pre-selecting a category."""

    def get(self, request, category_id=None):
        initial = {}
        if category_id:
            category = get_object_or_404(Category, pk=category_id)
            initial['category'] = category
            
        form = QuestionForm(initial=initial)
        return render(request, 'exams/admin/add_question.html', {
            'form': form,
            'category_id': category_id
        })

    def post(self, request, category_id=None):
        form = QuestionForm(request.POST)
        if form.is_valid():
            q = form.save(commit=False)
            q.created_by = request.user
            q.save()
            messages.success(request, 'Question added successfully.')
            
            # If we came from a category detail page, go back there
            if category_id:
                return redirect('admin_category_detail', pk=category_id)
            return redirect('admin_questions')
            
        return render(request, 'exams/admin/add_question.html', {
            'form': form,
            'category_id': category_id
        })


class AdminImportQuestionsView(BaseAdminRequiredMixin, View):
    """Bulk import questions, optionally forcing a specific category."""

    def get(self, request, category_id=None):
        category = None
        if category_id:
            category = get_object_or_404(Category, pk=category_id)
            
        form = ExcelImportForm()
        return render(request, 'exams/admin/import_questions.html', {
            'form': form,
            'category': category
        })

    def post(self, request, category_id=None):
        category = None
        if category_id:
            category = get_object_or_404(Category, pk=category_id)
            
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            result = import_questions_from_excel(
                request.FILES['excel_file'],
                created_by=request.user,
                forced_category=category
            )
            
            if result['created_count'] > 0:
                messages.success(
                    request,
                    f"Successfully imported {result['created_count']} questions."
                )
            if result['error_count'] > 0:
                messages.warning(
                    request,
                    f"{result['error_count']} rows had errors and were skipped."
                )
                
            return render(request, 'exams/admin/import_questions.html', {
                'form': ExcelImportForm(),
                'result': result,
                'category': category
            })
        return render(request, 'exams/admin/import_questions.html', {
            'form': form,
            'category': category
        })


class AdminQuestionDetailView(BaseAdminRequiredMixin, DetailView):
    """View a single question detail."""
    model = Question
    template_name = 'exams/admin/question_detail.html'
    context_object_name = 'question'


class AdminEditQuestionView(BaseAdminRequiredMixin, View):
    """Update an existing question."""

    def get(self, request, pk):
        question = get_object_or_404(Question, pk=pk)
        form = QuestionForm(instance=question)
        return render(request, 'exams/admin/add_question.html', {
            'form': form,
            'edit': True,
            'question': question
        })

    def post(self, request, pk):
        question = get_object_or_404(Question, pk=pk)
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            messages.success(request, 'Question updated successfully.')
            return redirect('admin_category_detail', pk=question.category.pk)
        return render(request, 'exams/admin/add_question.html', {
            'form': form,
            'edit': True,
            'question': question
        })


class AdminDeleteQuestionView(BaseAdminRequiredMixin, View):
    """Soft-delete a question."""

    def post(self, request, pk):
        question = get_object_or_404(Question, pk=pk)
        category_id = question.category.pk
        question.is_active = False
        question.save()
        messages.success(request, 'Question removed from question bank.')
        return redirect('admin_category_detail', pk=category_id)


class AdminBulkDeleteQuestionsView(BaseAdminRequiredMixin, View):
    """Bulk delete selected questions (soft-delete)."""

    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        category_id = request.POST.get('category_id')
        
        if not selected_ids:
            messages.warning(request, "No questions selected for deletion.")
            if category_id:
                return redirect('admin_category_detail', pk=category_id)
            return redirect('admin_dashboard')

        # Soft delete: set is_active=False
        count = Question.objects.filter(id__in=selected_ids).update(is_active=False)
        messages.success(request, f"Successfully removed {count} questions from the question bank.")
        
        if category_id:
            return redirect('admin_category_detail', pk=category_id)
        return redirect('admin_questions')


class AdminExamRequestsView(BaseAdminRequiredMixin, ListView):
    """Manage exam access requests — filter, approve, reject."""
    template_name = 'exams/admin/exam_requests.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        qs = ExamRequest.objects.select_related('student', 'category').order_by('-requested_at')
        status = self.request.GET.get('status', '').strip()
        if status:
            qs = qs.filter(status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['current_status'] = self.request.GET.get('status', '')
        ctx['status_choices'] = ExamRequest.Status.choices
        ctx['pending_count'] = ExamRequest.objects.filter(status='Pending').count()
        return ctx


class AdminApproveRequestView(BaseAdminRequiredMixin, View):
    """Approve a single exam request and generate paper."""

    def post(self, request, request_id):
        exam_req = get_object_or_404(ExamRequest, id=request_id, status='Pending')
        exam_req.approve(request.user)

        # Generate randomized paper for the student for the specific category
        try:
            paper = generate_exam_paper(
                student=exam_req.student, 
                category=exam_req.category
            )
            messages.success(
                request,
                f'Request approved and {exam_req.category.name} paper generated for {exam_req.student.username}.'
            )
        except ValueError as e:
            messages.warning(
                request,
                f'Request approved but paper generation failed: {e}. '
                'Please ensure enough questions exist in the selected category.'
            )

        return redirect('admin_exam_requests')


class AdminRejectRequestView(BaseAdminRequiredMixin, View):
    """Reject a single exam request with reason."""

    def post(self, request, request_id):
        exam_req = get_object_or_404(ExamRequest, id=request_id, status='Pending')
        reason = request.POST.get('rejection_reason', 'Rejected by admin').strip()
        exam_req.reject(request.user, reason=reason)
        messages.success(request, f'Request from {exam_req.student.username} rejected.')
        return redirect('admin_exam_requests')


class AdminBulkRequestActionView(BaseAdminRequiredMixin, View):
    """Bulk approve or reject multiple requests."""

    def post(self, request):
        action = request.POST.get('action', '')
        selected_ids = request.POST.getlist('selected_requests')

        if not selected_ids:
            messages.warning(request, 'No requests selected.')
            return redirect('admin_exam_requests')

        if action == 'approve':
            approved = ExamRequest.bulk_approve(selected_ids, request.user)
            # Generate papers for each approved student
            for req in approved:
                try:
                    generate_exam_paper(student=req.student, category=req.category)
                except ValueError:
                    pass
            messages.success(request, f'{len(approved)} requests approved.')

        elif action == 'reject':
            reason = request.POST.get('bulk_reason', 'Bulk rejection by admin')
            rejected = ExamRequest.bulk_reject(selected_ids, request.user, reason=reason)
            messages.success(request, f'{len(rejected)} requests rejected.')

        return redirect('admin_exam_requests')


class AdminResultsView(BaseAdminRequiredMixin, ListView):
    """View all exam results with analytics."""
    template_name = 'exams/admin/results.html'
    context_object_name = 'results'
    paginate_by = 20

    def get_filtered_queryset(self):
        qs = StudentExamResult.objects.select_related(
            'student', 'exam_paper', 'exam_paper__category'
        ).order_by('-started_at')

        status = self.request.GET.get('status', '').strip()
        search = self.request.GET.get('search', '').strip()
        min_pct = self.request.GET.get('min_percentage', '').strip()

        if status:
            qs = qs.filter(status=status)
        if search:
            qs = qs.filter(
                Q(student__username__icontains=search) |
                Q(student__first_name__icontains=search) |
                Q(student__last_name__icontains=search)
            )
        if min_pct:
            try:
                val = float(min_pct)
                from django.db.models.functions import Cast
                from django.db.models import FloatField
                qs = qs.filter(total_marks_possible__gt=0).annotate(
                    calc_pct=Cast(F('total_marks_obtained'), FloatField()) * 100 / Cast(F('total_marks_possible'), FloatField())
                ).filter(calc_pct__gte=val)
            except ValueError:
                pass
        return qs

    def get_queryset(self):
        return self.get_filtered_queryset()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['status_choices'] = StudentExamResult.Status.choices
        ctx['current_status'] = self.request.GET.get('status', '')
        ctx['search'] = self.request.GET.get('search', '')
        ctx['min_percentage'] = self.request.GET.get('min_percentage', '')

        # Summary stats based on filtered queryset
        qs = self.get_queryset()
        evaluated = qs.filter(status='Evaluated')
        ctx['total_evaluated'] = evaluated.count()
        avg = evaluated.aggregate(avg=Avg('total_marks_obtained'))
        ctx['avg_marks'] = round(avg['avg'] or 0, 1)

        return ctx


class AdminDeleteResultView(SuperuserRequiredMixin, View):
    """Delete a specific exam result."""
    def post(self, request, pk):
        result = get_object_or_404(StudentExamResult, pk=pk)
        student_name = result.student.username
        result.delete()
        messages.success(request, f"Exam result for {student_name} deleted successfully.")
        return redirect('admin_results')


class ExportResultsExcelView(AdminResultsView):
    """Export filtered exam results to Excel."""
    def get(self, request, *args, **kwargs):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, Alignment, PatternFill

        results = self.get_filtered_queryset()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exam Results"

        # Headers
        headers = ["Student", "Email", "Category", "Score", "Percentage", "Status", "Date"]
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="002147", end_color="002147", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        for r in results:
            ws.append([
                r.student.username,
                r.student.email or "—",
                r.exam_paper.category.name if r.exam_paper and r.exam_paper.category else "—",
                f"{r.total_marks_obtained} / {r.total_marks_possible}",
                f"{r.percentage()}%",
                r.status,
                r.started_at.strftime("%Y-%m-%d %H:%M")
            ])

        # Column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename=exam_results_{timezone.now().strftime('%Y%m%d')}.xlsx"
        wb.save(response)
        return response


class ExportResultsPDFView(AdminResultsView):
    """Export filtered exam results to PDF."""
    def get(self, request, *args, **kwargs):
        from django.template.loader import get_template
        from xhtml2pdf import pisa
        from django.http import HttpResponse

        results = self.get_filtered_queryset()
        template = get_template('exams/admin/export/results_pdf.html')
        context = {
            'results': results,
            'total_count': results.count(),
            'generated_at': timezone.now(),
            'search': self.request.GET.get('search', ''),
            'status': self.request.GET.get('status', ''),
        }
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="exam_results_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=500)
        return response


class AdminBulkDeleteResultsView(SuperuserRequiredMixin, View):
    """Bulk delete selected exam results."""
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            messages.warning(request, "No results selected for deletion.")
            return redirect('admin_results')
            
        count = StudentExamResult.objects.filter(id__in=selected_ids).delete()[0]
        messages.success(request, f"Successfully deleted {count} exam results.")
        return redirect('admin_results')


class AdminBulkDeleteStudentsView(SuperuserRequiredMixin, View):
    """Bulk delete selected students."""
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            messages.warning(request, "No students selected for deletion.")
            return redirect('admin_students')
            
        count = CustomUser.objects.filter(id__in=selected_ids, role='Student').delete()[0]
        messages.success(request, f"Successfully deleted {count} students.")
        return redirect('admin_students')


class AdminResultDetailView(BaseAdminRequiredMixin, DetailView):
    """Detailed result view — question-wise breakdown (admin only)."""
    template_name = 'exams/admin/result_detail.html'
    context_object_name = 'result'

    def get_queryset(self):
        return StudentExamResult.objects.select_related('student', 'exam_paper')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        result = self.object
        ctx['answers'] = result.answers.select_related('question').order_by('question')
        ctx['percentage'] = result.percentage()
        ctx['passed'] = result.is_passed()
        return ctx


class AdminNotificationsView(BaseAdminRequiredMixin, ListView):
    """Admin notification management."""
    template_name = 'exams/admin/notifications.html'
    context_object_name = 'notifications'
    paginate_by = 20

    def get_queryset(self):
        return Notification.objects.filter(
            recipient=self.request.user
        ).order_by('-created_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['unread_count'] = Notification.unread_count(self.request.user)
        return ctx


# ════════════════════════════════════════════════
#  ADMIN CATEGORY VIEWS
# ════════════════════════════════════════════════

class AdminCategoriesView(BaseAdminRequiredMixin, ListView):
    """List all exam categories."""
    template_name = 'exams/admin/categories.html'
    context_object_name = 'categories'
    
    def get_queryset(self):
        return Category.objects.annotate(
            q_count=Count('questions')
        ).order_by('name')


class AdminAddCategoryView(BaseAdminRequiredMixin, View):
    """Create a new exam category."""
    
    def get(self, request):
        form = CategoryForm()
        return render(request, 'exams/admin/add_category.html', {'form': form})
        
    def post(self, request):
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category created successfully.')
            return redirect('admin_categories')
        return render(request, 'exams/admin/add_category.html', {'form': form})


class AdminEditCategoryView(BaseAdminRequiredMixin, View):
    """Update an existing category."""
    
    def get(self, request, pk):
        cat = get_object_or_404(Category, pk=pk)
        form = CategoryForm(instance=cat)
        return render(request, 'exams/admin/add_category.html', {'form': form, 'edit': True})
        
    def post(self, request, pk):
        cat = get_object_or_404(Category, pk=pk)
        form = CategoryForm(request.POST, instance=cat)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated successfully.')
            return redirect('admin_categories')
        return render(request, 'exams/admin/add_category.html', {'form': form, 'edit': True})


class AdminDeleteCategoryView(BaseAdminRequiredMixin, View):
    """Delete a category (cascades to questions and exam papers)."""
    
    def post(self, request, pk):
        cat = get_object_or_404(Category, pk=pk)
        
        # Manually cascade to bypass PROTECT constraints
        # 1. Delete student answers for all questions in this category
        StudentAnswer.objects.filter(question__category=cat).delete()
        
        # 2. Delete student exam results for all papers in this category
        StudentExamResult.objects.filter(exam_paper__category=cat).delete()
        
        # 3. Delete exam papers in this category (cascades to ExamPaperQuestion)
        ExamPaper.objects.filter(category=cat).delete()
        
        # 4. Delete questions in this category
        Question.objects.filter(category=cat).delete()
        
        # 5. Delete category itself
        cat.delete()
        
        messages.success(request, 'Category and all related records deleted successfully.')
        return redirect('admin_categories')


class AdminBulkDeleteCategoriesView(BaseAdminRequiredMixin, View):
    """Bulk delete selected categories."""
    
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            messages.warning(request, "No categories selected for deletion.")
            return redirect('admin_categories')
            
        categories = Category.objects.filter(id__in=selected_ids)
        
        # Manually cascade for bulk delete
        StudentAnswer.objects.filter(question__category__in=categories).delete()
        StudentExamResult.objects.filter(exam_paper__category__in=categories).delete()
        ExamPaper.objects.filter(category__in=categories).delete()
        Question.objects.filter(category__in=categories).delete()
        
        count = categories.count()
        categories.delete()
        
        messages.success(request, f"Successfully deleted {count} categories and all related records.")
        return redirect('admin_categories')


class AdminToggleCategoryStatusView(BaseAdminRequiredMixin, View):
    """Toggle is_active status of a category."""
    
    def post(self, request, pk):
        cat = get_object_or_404(Category, pk=pk)
        cat.is_active = not cat.is_active
        cat.save(update_fields=['is_active'])
        status = 'Activated' if cat.is_active else 'Deactivated'
        messages.success(request, f'Category {cat.name} {status} successfully.')
        return redirect('admin_categories')


class PreviewCertificateView(View):
    """Generates and serves the certificate preview elegantly in HTML or directly as a PNG in the browser."""
    def get(self, request):
        from django.http import HttpResponse

        if request.GET.get('raw') != '1':
            html_content = """
            <!DOCTYPE html>
            <html>
            <head>
                <title>Certificate Preview</title>
                <style>
                    body, html {
                        margin: 0;
                        padding: 0;
                        width: 100%;
                        height: 100%;
                        background-color: #f4f4f4;
                        display: flex;
                        align-items: center;
                        justify-content: center;
                        overflow: hidden;
                    }
                    img {
                        max-width: 100vw;
                        max-height: 100vh;
                        object-fit: contain;
                        box-shadow: 0 4px 15px rgba(0,0,0,0.15);
                        background-color: #fff;
                    }
                </style>
            </head>
            <body>
                <img src="?raw=1" alt="Certificate Preview">
            </body>
            </html>
            """
            return HttpResponse(html_content)

        from .models import StudentExamResult, Category, ExamPaper, CustomUser
        import os
        from PIL import Image, ImageDraw, ImageFont
        from django.conf import settings
        from django.utils import timezone

        # Try to get the latest available exam result or create a dummy object
        result = StudentExamResult.objects.first()
        if not result:
            class MockCategory:
                name = "Python Programming"

            class MockExamPaper:
                category = MockCategory()

            class MockStudent:
                first_name = "Harish"
                last_name = "Subramanian"
                username = "harish"
                email = "test@example.com"
                profile_photo = None

            class MockResult:
                id = timezone.now()
                student = MockStudent()
                exam_paper = MockExamPaper()
                def percentage(self):
                    return 100.0

            result = MockResult()

        student = result.student
        student_name = f"{student.first_name} {student.last_name}".strip() or student.username
        course_name = result.exam_paper.category.name if result.exam_paper and result.exam_paper.category else "Aptitude Course"
        score = float(result.percentage()) if hasattr(result, 'percentage') else 100.0
        date_str = timezone.now().strftime("%B %Y")

        # Create a blank white canvas exactly A4 Landscape size (1414 x 1000)
        img = Image.new("RGB", (1414, 1000), color="#FFFFFF")
        draw = ImageDraw.Draw(img)

        # 1. Scale & Resolution (300 DPI: 3508 x 2480)
        WIDTH, HEIGHT = 3508, 2480
        SCALE = WIDTH / 1414.0 # Base scale relative to 1414 width
        
        img = Image.new("RGB", (WIDTH, HEIGHT), color="#FFFFFF")
        draw = ImageDraw.Draw(img)

        def find_font_file(font_name):
            if not font_name:
                return None

            if os.path.isabs(font_name) and os.path.exists(font_name):
                return font_name

            base_candidates = [
                os.path.join(str(settings.BASE_DIR), "static", "fonts", font_name),
                os.path.join(str(settings.BASE_DIR), "static", "images", font_name),
                os.path.join(str(settings.STATIC_ROOT), "fonts", font_name) if getattr(settings, "STATIC_ROOT", None) else None,
                os.path.join(str(settings.STATIC_ROOT), "images", font_name) if getattr(settings, "STATIC_ROOT", None) else None,
            ]
            base_candidates = [p for p in base_candidates if p]

            for candidate in base_candidates:
                if os.path.exists(candidate):
                    return candidate

            candidate_names = [font_name]
            lower = font_name.lower()
            if "times" in lower or "serif" in lower:
                candidate_names.extend([
                    "Times New Roman.ttf",
                    "Times.ttf",
                    "DejaVuSerif.ttf",
                    "DejaVuSerif-Bold.ttf",
                    "LiberationSerif-Regular.ttf",
                    "FreeSerif.ttf",
                ])
            elif "arial" in lower or "sans" in lower:
                candidate_names.extend([
                    "Arial.ttf",
                    "DejaVuSans.ttf",
                    "DejaVuSans-Bold.ttf",
                    "LiberationSans-Regular.ttf",
                    "LiberationSans-Bold.ttf",
                    "FreeSans.ttf",
                ])
            else:
                candidate_names.extend([
                    "DejaVuSans.ttf",
                    "DejaVuSerif.ttf",
                    "LiberationSans-Regular.ttf",
                    "LiberationSerif-Regular.ttf",
                    "FreeSans.ttf",
                    "FreeSerif.ttf",
                ])

            system_dirs = [
                os.path.join("C:", "Windows", "Fonts"),
                os.path.join("/usr", "share", "fonts"),
                os.path.join("/usr", "local", "share", "fonts"),
                os.path.expanduser("~/.fonts"),
                os.path.join("/Library", "Fonts"),
                os.path.join("/System", "Library", "Fonts"),
                os.path.join("/usr", "share", "fonts", "truetype"),
                os.path.join("/usr", "share", "fonts", "opentype"),
            ]

            for dir_path in system_dirs:
                if not dir_path or not os.path.isdir(dir_path):
                    continue
                for candidate in candidate_names:
                    candidate_path = os.path.join(dir_path, candidate)
                    if os.path.exists(candidate_path):
                        return candidate_path

            # Last resort: any valid TrueType/OpenType font on the host.
            for dir_path in system_dirs:
                if not os.path.isdir(dir_path):
                    continue
                for root, _, files in os.walk(dir_path):
                    for file_name in files:
                        if file_name.lower().endswith((".ttf", ".otf")):
                            candidate_path = os.path.join(root, file_name)
                            if os.path.exists(candidate_path):
                                return candidate_path

            return None

        def find_any_font_file():
            system_dirs = [
                os.path.join("C:", "Windows", "Fonts"),
                os.path.join("/usr", "share", "fonts"),
                os.path.join("/usr", "local", "share", "fonts"),
                os.path.expanduser("~/.fonts"),
                os.path.join("/Library", "Fonts"),
                os.path.join("/System", "Library", "Fonts"),
                os.path.join("/usr", "share", "fonts", "truetype"),
                os.path.join("/usr", "share", "fonts", "opentype"),
            ]
            for dir_path in system_dirs:
                if not os.path.isdir(dir_path):
                    continue
                for root, _, files in os.walk(dir_path):
                    for file_name in files:
                        if file_name.lower().endswith((".ttf", ".otf")):
                            candidate_path = os.path.join(root, file_name)
                            if os.path.exists(candidate_path):
                                return candidate_path
            return None

        def _apply_variation(font_obj, weight):
            """Set the weight axis on variable fonts; no-op on static fonts."""
            if weight is None or font_obj is None:
                return font_obj
            try:
                font_obj.set_variation_by_axes([int(weight)])
            except (AttributeError, OSError, ValueError, TypeError):
                pass
            return font_obj

        def get_font(font_name, size, weight=None):
            scaled_size = int(size * SCALE)
            font_path = find_font_file(font_name)
            if font_path:
                try:
                    return _apply_variation(
                        ImageFont.truetype(font_path, scaled_size), weight
                    )
                except OSError:
                    pass

            fallback_fonts = [
                "DejaVuSans-Bold.ttf",
                "DejaVuSans.ttf",
                "LiberationSans-Bold.ttf",
                "LiberationSans-Regular.ttf",
                "Arial.ttf",
                "FreeSans.ttf",
                "DejaVuSerif.ttf",
                "LiberationSerif-Regular.ttf",
            ]

            for path in fallback_fonts:
                try:
                    return _apply_variation(
                        ImageFont.truetype(path, scaled_size), weight
                    )
                except OSError:
                    continue

            any_font = find_any_font_file()
            if any_font:
                try:
                    return _apply_variation(
                        ImageFont.truetype(any_font, scaled_size), weight
                    )
                except OSError:
                    pass

            return ImageFont.load_default()

        def get_resampling_filter():
            if hasattr(Image, "Resampling"):
                return Image.Resampling.LANCZOS
            if hasattr(Image, "LANCZOS"):
                return Image.LANCZOS
            if hasattr(Image, "ANTIALIAS"):
                return Image.ANTIALIAS
            return Image.BICUBIC

        def text_width(text, font_obj):
            if hasattr(draw, "textlength"):
                try:
                    return draw.textlength(text, font=font_obj)
                except Exception:
                    pass
            if hasattr(draw, "textbbox"):
                bbox = draw.textbbox((0, 0), text, font=font_obj)
                return bbox[2] - bbox[0]
            if hasattr(font_obj, "getsize"):
                return font_obj.getsize(text)[0]
            return len(text) * getattr(font_obj, "size", 10)

        def resolve_asset_path(filename):
            candidate_dirs = [
                os.path.join(str(settings.BASE_DIR), 'static', 'images'),
            ]
            if getattr(settings, 'STATIC_ROOT', None):
                candidate_dirs.append(os.path.join(str(settings.STATIC_ROOT), 'images'))
            candidate_dirs = [d for d in candidate_dirs if d]
            for directory in candidate_dirs:
                candidate = os.path.join(directory, filename)
                if os.path.exists(candidate):
                    return candidate
            return None

        resample_filter = get_resampling_filter()

        # ─── Premium Certificate Typography (Editorial Luxury pairing) ──────
        # Title & Student Name : Playfair Display Bold      (elegant editorial serif)
        # Sub-headings & Course: Cormorant Garamond Bold/Medium (luxury serif)
        # Body & Footer        : Montserrat Regular/Medium  (clean modern sans)
        # Bundled variable fonts in static/fonts/; weight axis selected at render time.
        PLAYFAIR = "PlayfairDisplay-Variable.ttf"
        CORMORANT = "CormorantGaramond-Variable.ttf"
        MONTSERRAT = "Montserrat-Variable.ttf"

        font_cert    = get_font(PLAYFAIR,   85, weight=700)   # Certificate Title (Bold)
        font_sub     = get_font(CORMORANT,  32, weight=700)   # "OF ACHIEVEMENT" pill
        font_present = get_font(CORMORANT,  22, weight=500)   # Presenter Line (Medium)
        font_name    = get_font(PLAYFAIR,   78, weight=700)   # Student Name (Bold)
        font_course  = get_font(CORMORANT,  42, weight=700)   # Course banner (Bold)
        font_content = get_font(MONTSERRAT, 24, weight=400)   # Body description (Regular)
        font_footer  = get_font(MONTSERRAT, 20, weight=500)   # Footer (Medium)

        # Better wave implementation using polygons for accuracy
        import math
        import textwrap
        
        # Deepest Navy Wave (Base)
        wave_points = []
        for x in range(-100, WIDTH + 100, 10):
            # Scale the frequency and amplitude
            y = (280 * SCALE) + math.cos(x / (300.0 * SCALE)) * (100 * SCALE)
            wave_points.append((x, y))
        wave_points.extend([(WIDTH, 0), (0, 0)])
        draw.polygon(wave_points, fill="#000B1D")

        # Gold Wave Transition
        gold_points = []
        for x in range(-100, WIDTH + 100, 10):
            y = (310 * SCALE) + math.cos(x / (300.0 * SCALE)) * (100 * SCALE)
            gold_points.append((x, y))
        gold_points.extend([(WIDTH, 0), (0, 0)])
        draw.polygon(gold_points, fill="#C5A028")

        # Rich Navy Wave (Top)
        top_points = []
        for x in range(-100, WIDTH + 100, 10):
            y = (295 * SCALE) + math.cos(x / (300.0 * SCALE)) * (100 * SCALE)
            top_points.append((x, y))
        top_points.extend([(WIDTH, 0), (0, 0)])
        draw.polygon(top_points, fill="#001F3F")

        # Flowing Ribbons Top Right
        # Ribbon 1
        r1 = [(1100*SCALE, 0), (WIDTH, 150*SCALE), (WIDTH, 250*SCALE), (1100*SCALE, 50*SCALE)]
        draw.polygon(r1, fill="#D4AF37")
        # Ribbon 2
        r2 = [(1200*SCALE, 0), (WIDTH, 100*SCALE), (WIDTH, 180*SCALE), (1200*SCALE, 40*SCALE)]
        draw.polygon(r2, fill="#C5A028")

        # Path Definitions for Assets
        def resolve_asset_path(filename):
            candidates = [
                os.path.join(settings.BASE_DIR, 'static', 'images', filename),
                os.path.join(settings.BASE_DIR, 'staticfiles', 'images', filename),
            ]
            if getattr(settings, 'STATIC_ROOT', None):
                candidates.append(os.path.join(settings.STATIC_ROOT, 'images', filename))
            for candidate in candidates:
                if os.path.exists(candidate):
                    return candidate
            return None

        logo_path = resolve_asset_path('dlklogo.png')
        if not logo_path:
            logo_path = resolve_asset_path('dlklogo.jpg')
        if not logo_path:
            logo_path = resolve_asset_path('Logo.png')

        skill_path = resolve_asset_path('Skill India.png')
        iso_path = resolve_asset_path('Seal Image.png')
        sig_path = resolve_asset_path('Signature.png')

        # 3. Enhanced Design Elements
        # (Watermark removed for cleaner style)

        # 4. Logos with Refined Badge Backgrounds
        # DLK Logo (Top Left) - Enhanced for vibrancy and size
        if os.path.exists(logo_path):
            try:
                from PIL import ImageEnhance, ImageFilter
                logo = Image.open(logo_path).convert("RGBA")
                
                # 1. Enhance Quality (Increased for maximum visibility)
                brightness = ImageEnhance.Brightness(logo).enhance(1.4)
                contrast = ImageEnhance.Contrast(brightness).enhance(1.2)
                sharpness = ImageEnhance.Sharpness(contrast).enhance(2.0)
                vibrancy = ImageEnhance.Color(sharpness).enhance(1.6)
                logo = vibrancy
                
                # 2. Adjusted size for better balance
                logo.thumbnail((int(650*SCALE), int(300*SCALE)), resample_filter)
                
                lx = int(20*SCALE)
                ly = int(40*SCALE)
                
                # 3. Premium Soft White Glow effect
                glow_size = int(10 * SCALE)
                glow = Image.new("RGBA", (logo.width + glow_size*2, logo.height + glow_size*2), (0, 0, 0, 0))
                glow_mask = logo.split()[3]
                glow_silhouette = Image.new("RGBA", logo.size, (255, 255, 255, 70))
                glow.paste(glow_silhouette, (glow_size, glow_size), glow_mask)
                glow = glow.filter(ImageFilter.GaussianBlur(glow_size))
                img.paste(glow, (lx - glow_size, ly - glow_size), glow)
                
                img.paste(logo, (lx, ly), logo if logo.mode == 'RGBA' else None)
            except Exception:
                pass

        # Skill India (Top Right)
        if os.path.exists(skill_path):
            try:
                skill_img = Image.open(skill_path).convert("RGBA")
                # Increase size
                skill_img.thumbnail((int(400*SCALE), int(200*SCALE)), resample_filter)
                sx = int(WIDTH - skill_img.width - 40*SCALE) # Match left margin
                sy = int(40*SCALE)
                img.paste(skill_img, (sx, sy), skill_img)
            except Exception:
                pass

        # 5. Main Text Content
        # "CERTIFICATE" (with deep shadow)
        draw.text((707*SCALE + 3*SCALE, 120*SCALE + 3*SCALE), "CERTIFICATE", font=font_cert, fill="#00000044", anchor="mm")
        draw.text((707*SCALE, 120*SCALE), "CERTIFICATE", font=font_cert, fill="#C5A028", anchor="mm") # Gold Title
        
        # "OF ACHIEVEMENT" Pill
        pill_coords = [(480*SCALE, 185*SCALE), (934*SCALE, 245*SCALE)]
        draw.rounded_rectangle(pill_coords, radius=30*SCALE, fill="#003366", outline="#D4AF37", width=int(3*SCALE))
        draw.text((707*SCALE, 215*SCALE), "OF ACHIEVEMENT", font=font_sub, fill="#FFFFFF", anchor="mm")

        # "THIS CERTIFICATE IS PROUDLY PRESENT TO"
        draw.text((707*SCALE, 430*SCALE), "THIS CERTIFICATE IS PROUDLY PRESENT TO", font=font_present, fill="#555555", anchor="mm")
        
        # Student Name (Luxury Rendering with Letter Spacing)
        name_upper = student_name.upper()
        
        def draw_text_spaced(draw_obj, position, text, font, spacing, fill, anchor_centered=True, stroke_width=0, stroke_fill=None):
            total_width = 0
            char_widths = []
            for char in text:
                cw = text_width(char, font)
                char_widths.append(cw)
                total_width += cw + (spacing if char != text[-1] else 0)
            
            start_x = position[0] - (total_width / 2) if anchor_centered else position[0]
            curr_x = start_x
            for i, char in enumerate(text):
                draw_obj.text((curr_x, position[1]), char, font=font, fill=fill, anchor="lm", stroke_width=stroke_width, stroke_fill=stroke_fill)
                curr_x += char_widths[i] + spacing

        def fit_name_font_and_spacing(text, font, max_width):
            base_spacing = int(8 * SCALE)
            min_spacing = int(2 * SCALE)

            def total_text_width(font_obj, spacing):
                return text_width(text, font_obj) + spacing * max(0, len(text) - 1)

            spacing = base_spacing
            while spacing >= min_spacing:
                if total_text_width(font, spacing) <= max_width:
                    return font, spacing
                spacing -= int(1 * SCALE)

            for font_size in range(78, int(78 * 0.55), -2):
                temp_font = get_font(PLAYFAIR, font_size, weight=700)
                if total_text_width(temp_font, min_spacing) <= max_width:
                    return temp_font, min_spacing

            return font, min_spacing

        spacing_val = int(8 * SCALE)
        # Keep the name within the center area and avoid overlapping the ISO seal on the right.
        center_x = 707 * SCALE
        seal_left = (1220 * SCALE) - (120 * SCALE)
        reserved_margin = 100 * SCALE
        max_name_width = int((seal_left - reserved_margin - center_x) * 2)
        font_name, spacing_val = fit_name_font_and_spacing(name_upper, font_name, max_name_width)
        
        draw_text_spaced(draw, (707*SCALE + 5, 520*SCALE + 5), name_upper, font_name, spacing_val, "#00000011")
        draw_text_spaced(draw, (707*SCALE + 2, 520*SCALE + 2), name_upper, font_name, spacing_val, "#C5A028")
        draw_text_spaced(draw, (707*SCALE, 520*SCALE), name_upper, font_name, spacing_val, "#000B1D", stroke_width=int(1*SCALE), stroke_fill="#000B1D")

        # Course Name Section (Dynamic Width Fitting)
        course_text_upper = course_name.upper()
        tw, th = draw.textbbox((0, 0), course_text_upper, font=font_course)[2:]
        
        px, py = 50 * SCALE, 15 * SCALE
        center_x = 707 * SCALE
        center_y = 605 * SCALE
        
        course_banner = [
            (center_x - (tw/2) - px, center_y - (th/2) - py),
            (center_x + (tw/2) + px, center_y + (th/2) + py)
        ]
        
        draw.rectangle(course_banner, fill="#C5A028") 
        inner_banner = [
            (course_banner[0][0] + 3*SCALE, course_banner[0][1] + 3*SCALE),
            (course_banner[1][0] - 3*SCALE, course_banner[1][1] - 3*SCALE)
        ]
        draw.rectangle(inner_banner, outline="#001F3F", width=int(2*SCALE))
        
        draw.text((center_x, center_y), course_text_upper, font=font_course, fill="#000B1D", anchor="mm", stroke_width=int(1*SCALE), stroke_fill="#000B1D")

        # Content Text
        aptitude_text = (
            "This certificate is awarded for successfully completing the Aptitude and Reasoning assessment. "
            "The candidate has demonstrated proficiency in logical reasoning, quantitative analysis, "
            "and problem-solving capabilities essential for professional excellence."
        )
        lines = textwrap.wrap(aptitude_text, width=80)
        y_text = 710 * SCALE
        for line in lines:
            draw.text((707*SCALE, y_text), line, font=font_content, fill="#444444", anchor="mm")
            y_text += 42 * SCALE

        # Website
       
        # 6. Signature & Date Section
        if os.path.exists(sig_path):
            try:
                from PIL import ImageEnhance
                sig_img = Image.open(sig_path).convert("RGBA")
                # Increase size further
                sig_img.thumbnail((int(750*SCALE), int(350*SCALE)), resample_filter)
                
                # Make it significantly darker, bolder and more visible (Ink effect)
                contrast_enhancer = ImageEnhance.Contrast(sig_img)
                sig_img = contrast_enhancer.enhance(3.0) 
                
                brightness_enhancer = ImageEnhance.Brightness(sig_img)
                sig_img = brightness_enhancer.enhance(0.4) 
                
                sharpness_enhancer = ImageEnhance.Sharpness(sig_img)
                sig_img = sharpness_enhancer.enhance(2.5) 
                
                sx = int(300*SCALE - (sig_img.width // 2))
                # Move to the absolute bottom, almost touching the border line
                sy = int(1080*SCALE - sig_img.height)
                img.paste(sig_img, (sx, sy), sig_img)
            except Exception:
                pass
        
        draw.line([(100*SCALE, 920*SCALE), (500*SCALE, 920*SCALE)], fill="#002147", width=int(1*SCALE))
        draw.text((300*SCALE, 940*SCALE), "SIGNATURE", font=font_footer, fill="#002147", anchor="mm")
        
        date_now = timezone.now().strftime("%d-%m-%Y")
        draw.line([(914*SCALE, 920*SCALE), (1314*SCALE, 920*SCALE)], fill="#002147", width=int(1*SCALE))
        draw.text((1114*SCALE, 940*SCALE), f"DATE: {date_now}", font=font_footer, fill="#002147", anchor="mm")

        # 7. Fixed ISO Seal & Professional Badge
        if os.path.exists(iso_path):
            try:
                # Draw professional ribbon with notched bottom
                ribbon_pts = [(1140*SCALE, 300*SCALE), (1300*SCALE, 300*SCALE), (1300*SCALE, 600*SCALE), (1220*SCALE, 550*SCALE), (1140*SCALE, 600*SCALE)]
                draw.polygon(ribbon_pts, fill="#000B1D") # Deeper Navy
                draw.polygon(ribbon_pts, outline="#C5A028", width=int(2*SCALE))
                
                # Draw the "Yellow Circular Badge" base
                seal_center = (1220*SCALE, 420*SCALE)
                seal_radius = 120*SCALE
                # Outer gold circle
                draw.ellipse([(seal_center[0]-seal_radius, seal_center[1]-seal_radius), (seal_center[0]+seal_radius, seal_center[1]+seal_radius)], fill="#C5A028", outline="#8B7500", width=int(3*SCALE))
                # Inner gold ring
                draw.ellipse([(seal_center[0]-seal_radius+8*SCALE, seal_center[1]-seal_radius+8*SCALE), (seal_center[0]+seal_radius-8*SCALE, seal_center[1]+seal_radius-8*SCALE)], outline="#F9E79F", width=int(2*SCALE))
                
                iso_img = Image.open(iso_path).convert("RGBA")
                # 1. Remove transparent padding
                bbox = iso_img.getbbox()
                if bbox:
                    iso_img = iso_img.crop(bbox)
                
                # 2. Force to Square to prevent distortion
                w, h = iso_img.size
                sq_size = max(w, h)
                square_iso = Image.new("RGBA", (sq_size, sq_size), (0, 0, 0, 0))
                square_iso.paste(iso_img, ((sq_size - w) // 2, (sq_size - h) // 2))
                
                # 3. Resize to fit perfectly inside the yellow circle without overflowing
                # Using 2.8x multiplier to ensure the seal stays within the gold circle boundaries
                iso_final_size = int(seal_radius * 2.8)
                iso_img_final = square_iso.resize((iso_final_size, iso_final_size), resample_filter)
                
                # 4. Perfect Centering with corrective nudge (adjusted for perceived center and user feedback)
                # Compensation for asset asymmetry and custom positioning
                nudge_x = 36 * SCALE
                nudge_y = -4 * SCALE
                paste_x = int(round(seal_center[0] - (iso_final_size / 2.0) - nudge_x))
                paste_y = int(round(seal_center[1] - (iso_final_size / 2.0) - nudge_y))
                img.paste(iso_img_final, (paste_x, paste_y), iso_img_final)
            except Exception:
                pass

        # Final 3px Border
        draw.rectangle([(0, 0), (WIDTH-1, HEIGHT-1)], outline="#C5A028", width=int(3*SCALE))

        # Return as PNG direct response
        response = HttpResponse(content_type="image/png")
        img.save(response, "PNG")
        return response



# ════════════════════════════════════════════════
#  NEW ANALYTICS & PERMISSION VIEWS
# ════════════════════════════════════════════════

class AdminNeverAttendedStudentsView(BaseAdminRequiredMixin, ListView):
    """View list of students who registered but never took an exam."""
    template_name = 'exams/admin/never_attended_students.html'
    context_object_name = 'students'
    paginate_by = 20

    def get_queryset(self):
        attended_student_ids = StudentExamResult.objects.values_list('student_id', flat=True).distinct()
        return CustomUser.objects.filter(role='Student').exclude(id__in=attended_student_ids).order_by('-date_joined')


class AdminStudentCertificateHistoryView(BaseAdminRequiredMixin, DetailView):
    """Full certificate history for a specific student."""
    template_name = 'exams/admin/student_certificate_history.html'
    context_object_name = 'student'
    model = CustomUser

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        student = self.object
        # Certificates are stored in the Certificate model (unique per category)
        ctx['certificates'] = Certificate.objects.filter(student=student).select_related('category').order_by('-updated_at')
        return ctx


# ════════════════════════════════════════════════
#  CUSTOM ERROR HANDLERS
# ════════════════════════════════════════════════

def error_404(request, exception):
    """Page Not Found"""
    return render(request, 'errors/404.html', status=404)

def error_500(request):
    """Internal Server Error"""
    return render(request, 'errors/500.html', status=500)

def error_403(request, exception=None):
    """Permission Denied"""
    return render(request, 'errors/403.html', status=403)

def error_400(request, exception=None):
    """Bad Request"""
    return render(request, 'errors/400.html', status=400)

def error_502(request):
    """Bad Gateway"""
    return render(request, 'errors/502.html', status=502)

def error_503(request):
    """Service Unavailable"""
    return render(request, 'errors/503.html', status=503)

