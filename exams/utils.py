"""
Utility functions for the exam system.

Includes:
    - Excel bulk import for questions
    - Randomized paper generation
    - Auto-evaluation runner
"""

import random
from django.db import transaction
from django.utils import timezone


def import_questions_from_excel(file_obj, created_by=None, forced_category=None):
    """
    Parse an uploaded Excel file and bulk-import questions.
    """
    try:
        import openpyxl
    except ImportError:
        return {
            'created_count': 0,
            'error_count': 1,
            'errors': [{'row': 0, 'error': 'openpyxl is not installed. Run: pip install openpyxl'}],
        }

    from .models import Question

    wb = openpyxl.load_workbook(file_obj, read_only=True)
    ws = wb.active

    # Extract headers from first row
    headers = []
    for cell in ws[1]:
        val = cell.value
        if val:
            headers.append(str(val).strip().lower().replace(' ', '_'))
        else:
            headers.append('')

    # Build rows as list of dicts
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = val
        # Skip completely empty rows
        if any(v is not None and str(v).strip() for v in row_dict.values()):
            rows.append(row_dict)

    wb.close()

    if not rows:
        return {
            'created_count': 0,
            'error_count': 0,
            'errors': [],
        }

    created, errors = Question.objects.bulk_import_from_rows(
        rows, created_by=created_by, forced_category=forced_category
    )

    return {
        'created_count': len(created),
        'error_count': len(errors),
        'errors': errors,
    }


def generate_exam_paper(student, category, num_questions=None):
    """
    Generate a randomized exam paper for a student for a specific category.

    Args:
        student: CustomUser instance (role=Student)
        category: Category instance
        num_questions: Optional number of questions to include. 
                       If None, all active questions in the category are included.

    Returns:
        ExamPaper instance with linked questions

    Raises:
        ValueError: If no questions available in the category
    """
    from .models import Question, ExamPaper, ExamPaperQuestion

    # Fetch all active questions for this category
    all_questions = list(
        Question.objects.filter(
            category=category,
            is_active=True,
        ).values_list('id', flat=True)
    )

    if not all_questions:
        raise ValueError(
            f'No active questions found in the "{category.name}" category. '
            'Please add questions before approving requests.'
        )

    # Determine how many questions to pick
    if num_questions and len(all_questions) > num_questions:
        selected_questions = random.sample(all_questions, num_questions)
    else:
        selected_questions = all_questions

    # Shuffle the final selection for random order
    random.shuffle(selected_questions)

    with transaction.atomic():
        # Prevent multiple active papers for the same student/category?
        # Actually, the view handles that by checking for existing result/paper.
        paper = ExamPaper.objects.create(
            student=student, 
            category=category
        )

        # Create through-model entries preserving order
        through_objects = [
            ExamPaperQuestion(
                exam_paper=paper,
                question_id=q_id,
                order=idx,
            )
            for idx, q_id in enumerate(selected_questions, start=1)
        ]
        ExamPaperQuestion.objects.bulk_create(through_objects, batch_size=500)

        # Recalculate totals from linked questions
        paper.recalculate_totals()

    return paper


def submit_and_evaluate(result):
    """
    Finalize an exam submission and run auto-evaluation.

    Args:
        result: StudentExamResult instance (status=In Progress)

    Returns:
        The updated StudentExamResult instance
    """
    from .models import StudentExamResult, Notification

    result.submitted_at = timezone.now()
    result.status = StudentExamResult.Status.SUBMITTED
    result.save(update_fields=['submitted_at', 'status'])

    # Auto-evaluate MCQ and True/False answers
    result.auto_evaluate()

    # Notify admin that a submission is ready for review (for essay/short answer)
    has_manual_review = result.answers.filter(
        question__question_type__in=['Short Answer', 'Essay']
    ).exists()

    if has_manual_review:
        # Notify admins (superusers) about manual review needed
        from .models import CustomUser
        admins = CustomUser.objects.filter(is_superuser=True, is_active=True)
        notifications = [
            Notification(
                recipient=admin,
                notification_type=Notification.NotificationType.INFO,
                title='Manual Review Required',
                message=f'{result.student.username} submitted an exam with questions requiring manual review.',
            )
            for admin in admins
        ]
        Notification.objects.bulk_create(notifications, batch_size=100)

    return result


def generate_and_send_certificate(result):
    """
    Generate the certificate using Pillow in A4 Landscape style and send via email.
    """
    import os
    import tempfile
    from PIL import Image, ImageDraw, ImageFont
    from django.conf import settings
    from django.core.mail import EmailMessage
    from django.utils import timezone

    from .models import Certificate
    
    student = result.student
    student_name = f"{student.first_name} {student.last_name}".strip() or student.username
    category = result.exam_paper.category
    course_name = category.name if category else "Aptitude Course"
    score = float(result.percentage())
    date_str = timezone.now().strftime("%B %Y")

    # 1. Update or Create Certificate record
    # Enforces "One student + One course/category = Only ONE certificate record"
    certificate, created = Certificate.objects.update_or_create(
        student=student,
        category=category,
        defaults={
            'marks': result.total_marks_obtained,
            'percentage': result.percentage(),
            'total_marks': result.total_marks_possible,
            'exam_date': result.completed_at or timezone.now(),
        }
    )

    # 2. Scale & Resolution (300 DPI: 3508 x 2480)
    WIDTH, HEIGHT = 3508, 2480
    SCALE = WIDTH / 1414.0 # Base scale relative to 1414 width
    
    img = Image.new("RGB", (WIDTH, HEIGHT), color="#FFFFFF")
    draw = ImageDraw.Draw(img)

    def get_font(font_name, size):
        scaled_size = int(size * SCALE)
        # Potential font paths on Linux
        linux_font_paths = [
            "/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSerif.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        ]
        
        # 1. Try requested font
        try:
            return ImageFont.truetype(font_name, scaled_size)
        except OSError:
            pass
            
        # 2. Try common system fonts
        for path in ["arial.ttf", "times.ttf", "Arial.ttf", "Times.ttf"] + linux_font_paths:
            try:
                return ImageFont.truetype(path, scaled_size)
            except OSError:
                continue
                
        # 3. Last resort (will be small)
        return ImageFont.load_default()

    # Scaled Fonts
    font_cert = get_font("times.ttf", 85)
    font_sub = get_font("arial.ttf", 32)
    font_present = get_font("arial.ttf", 22)
    font_name = get_font("times.ttf", 78) # Decreased 2pts as requested
    font_course = get_font("times.ttf", 42) # Changed to elegant serif (Times)
    font_content = get_font("arial.ttf", 24)
    font_footer = get_font("arial.ttf", 20)

    # Better wave implementation using polygons for accuracy
    import math
    import textwrap
    
    # Deepest Navy Wave (Base)
    wave_points = []
    for x in range(-100, WIDTH + 100, 10):
        y = (280 * SCALE) + math.cos(x / (300.0 * SCALE)) * (100 * SCALE)
        wave_points.append((x, y))
    wave_points.extend([(WIDTH, 0), (0, 0)])
    draw.polygon(wave_points, fill="#000B1D")

    # Gold Wave Transition
    gold_points = []
    for x in range(-100, WIDTH + 100, 10):
        y = (310 * SCALE) + math.cos(x / (300.0 * SCALE)) * (100 * SCALE)
        gold_points.append((x, y))
    gold_points.extend([(WIDTH, 0), (0, 0)])
    draw.polygon(gold_points, fill="#C5A028")

    # Rich Navy Wave (Top)
    top_points = []
    for x in range(-100, WIDTH + 100, 10):
        y = (295 * SCALE) + math.cos(x / (300.0 * SCALE)) * (100 * SCALE)
        top_points.append((x, y))
    top_points.extend([(WIDTH, 0), (0, 0)])
    draw.polygon(top_points, fill="#001F3F")

    # Path Definitions for Assets
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'dlklogo.png')
    if not os.path.exists(logo_path):
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'dlklogo.jpg')
    if not os.path.exists(logo_path):
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'Logo.png')
        
    skill_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'Skill India.png')
    iso_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'Seal Image.png')
    sig_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'Signature.png')

    # 3. Enhanced Design Elements
    # (Watermark removed for cleaner style)

    # 4. Logos with Refined Badge Backgrounds
    # DLK Logo (Top Left) - Enhanced for vibrancy and size
    if os.path.exists(logo_path):
        try:
            from PIL import ImageEnhance, ImageFilter
            logo = Image.open(logo_path).convert("RGBA")
            
            # 1. Enhance Quality (Increased for maximum visibility)
            brightness = ImageEnhance.Brightness(logo).enhance(1.4) # +40% Brightness
            contrast = ImageEnhance.Contrast(brightness).enhance(1.2) # +20% Contrast
            sharpness = ImageEnhance.Sharpness(contrast).enhance(2.0) # +100% Sharpness
            vibrancy = ImageEnhance.Color(sharpness).enhance(1.6) # +60% Saturation (Green focus)
            logo = vibrancy
            
            # 2. Adjusted size for better balance
            logo.thumbnail((int(650*SCALE), int(300*SCALE)), Image.Resampling.LANCZOS)
            
            lx = int(20*SCALE)
            ly = int(40*SCALE)
            
            # 3. Premium Soft White Glow effect
            glow_size = int(10 * SCALE)
            glow = Image.new("RGBA", (logo.width + glow_size*2, logo.height + glow_size*2), (0, 0, 0, 0))
            # Paste a soft white silhouette for glow
            glow_mask = logo.split()[3]
            glow_silhouette = Image.new("RGBA", logo.size, (255, 255, 255, 70)) # Soft White
            glow.paste(glow_silhouette, (glow_size, glow_size), glow_mask)
            glow = glow.filter(ImageFilter.GaussianBlur(glow_size))
            img.paste(glow, (lx - glow_size, ly - glow_size), glow)
            
            img.paste(logo, (lx, ly), logo if logo.mode == 'RGBA' else None)
        except Exception:
            pass

    # Skill India (Top Right)
    if os.path.exists(skill_path):
        try:
            skill_img = Image.open(skill_path).convert("RGBA")
            # Increase size
            skill_img.thumbnail((int(400*SCALE), int(200*SCALE)), Image.Resampling.LANCZOS)
            sx = int(WIDTH - skill_img.width - 40*SCALE) # Match left margin
            sy = int(40*SCALE)
            img.paste(skill_img, (sx, sy), skill_img)
        except Exception:
            pass

    # 5. Main Text Content
    # "CERTIFICATE" (with deep shadow)
    draw.text((707*SCALE + 3*SCALE, 120*SCALE + 3*SCALE), "CERTIFICATE", font=font_cert, fill="#00000044", anchor="mm")
    draw.text((707*SCALE, 120*SCALE), "CERTIFICATE", font=font_cert, fill="#C5A028", anchor="mm") # Gold Title
    
    # "OF ACHIEVEMENT" Pill
    pill_coords = [(480*SCALE, 185*SCALE), (934*SCALE, 245*SCALE)]
    draw.rounded_rectangle(pill_coords, radius=30*SCALE, fill="#003366", outline="#D4AF37", width=int(3*SCALE))
    draw.text((707*SCALE, 215*SCALE), "OF ACHIEVEMENT", font=font_sub, fill="#FFFFFF", anchor="mm")

    # "THIS CERTIFICATE IS PROUDLY PRESENT TO"
    draw.text((707*SCALE, 430*SCALE), "THIS CERTIFICATE IS PROUDLY PRESENT TO", font=font_present, fill="#555555", anchor="mm")
    
    # Student Name (Luxury Rendering with Letter Spacing)
    name_upper = student_name.upper()
    
    # Custom function for character spacing (Letter Spacing)
    def draw_text_spaced(draw_obj, position, text, font, spacing, fill, anchor_centered=True, stroke_width=0, stroke_fill=None):
        # Calculate total width first
        total_width = 0
        char_widths = []
        for char in text:
            cw = draw_obj.textlength(char, font=font)
            char_widths.append(cw)
            total_width += cw + (spacing if char != text[-1] else 0)
        
        start_x = position[0] - (total_width / 2) if anchor_centered else position[0]
        curr_x = start_x
        for i, char in enumerate(text):
            draw_obj.text((curr_x, position[1]), char, font=font, fill=fill, anchor="lm", stroke_width=stroke_width, stroke_fill=stroke_fill)
            curr_x += char_widths[i] + spacing

    spacing_val = int(8 * SCALE) # Elegant spacing
    
    # Draw Depth Layers
    # 1. Soft broad shadow
    draw_text_spaced(draw, (707*SCALE + 5*SCALE, 520*SCALE + 5*SCALE), name_upper, font_name, spacing_val, "#00000011")
    # 2. Tight gold accent
    draw_text_spaced(draw, (707*SCALE + 2*SCALE, 520*SCALE + 2*SCALE), name_upper, font_name, spacing_val, "#C5A028")
    # 3. Main Luxury Text with Sharp Stroke
    draw_text_spaced(draw, (707*SCALE, 520*SCALE), name_upper, font_name, spacing_val, "#000B1D", stroke_width=int(1*SCALE), stroke_fill="#000B1D")

    # Course Name Section (Dynamic Width Fitting)
    # 1. Calculate precise text dimensions
    course_text_upper = course_name.upper()
    tw, th = draw.textbbox((0, 0), course_text_upper, font=font_course)[2:]
    
    # 2. Define dynamic banner with elegant padding
    px, py = 50 * SCALE, 15 * SCALE
    center_x = 707 * SCALE
    center_y = 605 * SCALE
    
    course_banner = [
        (center_x - (tw/2) - px, center_y - (th/2) - py),
        (center_x + (tw/2) + px, center_y + (th/2) + py)
    ]
    
    # 3. Premium Gold Gradient Background Styling
    draw.rectangle(course_banner, fill="#C5A028") # Royal Gold Base
    # Subtle inner navy highlight for "gradient" effect
    inner_banner = [
        (course_banner[0][0] + 3*SCALE, course_banner[0][1] + 3*SCALE),
        (course_banner[1][0] - 3*SCALE, course_banner[1][1] - 3*SCALE)
    ]
    draw.rectangle(inner_banner, outline="#001F3F", width=int(2*SCALE))
    
    draw.text((center_x, center_y), course_text_upper, font=font_course, fill="#000B1D", anchor="mm", stroke_width=int(1*SCALE), stroke_fill="#000B1D") # Added stroke for boldness

    # Content Text
    aptitude_text = (
        "This certificate is awarded for successfully completing the Aptitude and Reasoning assessment. "
        "The candidate has demonstrated proficiency in logical reasoning, quantitative analysis, "
        "and problem-solving capabilities essential for professional excellence."
    )
    import textwrap
    lines = textwrap.wrap(aptitude_text, width=80)
    y_text = 710 * SCALE
    for line in lines:
        draw.text((707*SCALE, y_text), line, font=font_content, fill="#444444", anchor="mm")
        y_text += 42 * SCALE

    # Website

    # 6. Signature & Date Section
    if os.path.exists(sig_path):
        try:
            from PIL import ImageEnhance
            sig_img = Image.open(sig_path).convert("RGBA")
            # Increase size further
            sig_img.thumbnail((int(750*SCALE), int(350*SCALE)), Image.Resampling.LANCZOS)
            
            # Make it significantly darker, bolder and more visible (Ink effect)
            contrast_enhancer = ImageEnhance.Contrast(sig_img)
            sig_img = contrast_enhancer.enhance(3.0)  # Extreme contrast for ink depth
            
            brightness_enhancer = ImageEnhance.Brightness(sig_img)
            sig_img = brightness_enhancer.enhance(0.4)  # Significantly darker for bold ink look
            
            sharpness_enhancer = ImageEnhance.Sharpness(sig_img)
            sig_img = sharpness_enhancer.enhance(2.5)  # Crisper edges
            
            sx = int(300*SCALE - (sig_img.width // 2))
            # Move to the absolute bottom, almost touching the border line
            sy = int(1080*SCALE - sig_img.height)
            img.paste(sig_img, (sx, sy), sig_img)
        except Exception:
            pass
    
    draw.line([(100*SCALE, 920*SCALE), (500*SCALE, 920*SCALE)], fill="#002147", width=int(1*SCALE))
    draw.text((300*SCALE, 940*SCALE), "SIGNATURE", font=font_footer, fill="#002147", anchor="mm")
    
    date_now = timezone.now().strftime("%d-%m-%Y")
    draw.line([(914*SCALE, 920*SCALE), (1314*SCALE, 920*SCALE)], fill="#002147", width=int(1*SCALE))
    draw.text((1114*SCALE, 940*SCALE), f"DATE: {date_now}", font=font_footer, fill="#002147", anchor="mm")

    # 7. Fixed ISO Seal & Professional Ribbon
    if os.path.exists(iso_path):
        try:
            # Draw professional ribbon with notched bottom
            ribbon_pts = [(1140*SCALE, 300*SCALE), (1300*SCALE, 300*SCALE), (1300*SCALE, 600*SCALE), (1220*SCALE, 550*SCALE), (1140*SCALE, 600*SCALE)]
            draw.polygon(ribbon_pts, fill="#000B1D") # Deeper Navy
            draw.polygon(ribbon_pts, outline="#C5A028", width=int(2*SCALE))
            
            # Draw the "Yellow Circular Badge" base
            seal_center = (1220*SCALE, 420*SCALE)
            seal_radius = 120*SCALE
            # Outer gold circle
            draw.ellipse([(seal_center[0]-seal_radius, seal_center[1]-seal_radius), (seal_center[0]+seal_radius, seal_center[1]+seal_radius)], fill="#C5A028", outline="#8B7500", width=int(3*SCALE))
            # Inner gold ring
            draw.ellipse([(seal_center[0]-seal_radius+8*SCALE, seal_center[1]-seal_radius+8*SCALE), (seal_center[0]+seal_radius-8*SCALE, seal_center[1]+seal_radius-8*SCALE)], outline="#F9E79F", width=int(2*SCALE))
            
            iso_img = Image.open(iso_path).convert("RGBA")
            # 1. Remove transparent padding
            bbox = iso_img.getbbox()
            if bbox:
                iso_img = iso_img.crop(bbox)
            
            # 2. Force to Square to prevent distortion
            w, h = iso_img.size
            sq_size = max(w, h)
            square_iso = Image.new("RGBA", (sq_size, sq_size), (0, 0, 0, 0))
            square_iso.paste(iso_img, ((sq_size - w) // 2, (sq_size - h) // 2))
            
            # 3. Resize to fit perfectly inside the yellow circle without overflowing
            # Using 2.8x multiplier to ensure the seal stays within the gold circle boundaries
            iso_final_size = int(seal_radius * 2.8)
            iso_img_final = square_iso.resize((iso_final_size, iso_final_size), Image.Resampling.LANCZOS)
            
            # 4. Perfect Centering with corrective nudge (adjusted for perceived center and user feedback)
            # Compensation for asset asymmetry and custom positioning
            nudge_x = 36 * SCALE
            nudge_y = -4 * SCALE
            paste_x = int(round(seal_center[0] - (iso_final_size / 2.0) - nudge_x))
            paste_y = int(round(seal_center[1] - (iso_final_size / 2.0) - nudge_y))
            img.paste(iso_img_final, (paste_x, paste_y), iso_img_final)
        except Exception:
            pass

    # Final 3px Border
    draw.rectangle([(0, 0), (WIDTH-1, HEIGHT-1)], outline="#C5A028", width=int(3*SCALE))

    # 8. Save and send via email
    cert_filename = f"Certificate_{student.username}_{result.id.hex[:6]}.png"
    cert_path = os.path.join(tempfile.gettempdir(), cert_filename)
    img.save(cert_path, "PNG")

    subject = f"Congratulations {student.first_name or student.username}! Your Certificate of Completion"
    body = f"""Dear {student.first_name or student.username},

Congratulations! You have successfully passed the assessment for {course_name} with a score of {score:.2f}%.

Please find your certificate attached to this email.

Best regards,
Aptipro Exam Team"""

    email_msg = EmailMessage(
        subject=subject,
        body=body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[student.email],
    )
    email_msg.attach_file(cert_path)
    email_msg.send(fail_silently=False)


